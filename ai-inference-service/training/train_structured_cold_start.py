from __future__ import annotations

import argparse
import hashlib
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path

import numpy as np
import onnx
import onnxruntime as ort
from onnx import TensorProto, helper, numpy_helper
from openpyxl import load_workbook
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, balanced_accuracy_score, confusion_matrix
from sklearn.model_selection import GroupShuffleSplit
from sklearn.preprocessing import StandardScaler


RANDOM_STATE = 42
MODEL_CODE = "structured-classifier"
MODEL_VERSION = "0.1.0-warning-weak-label"
TARGETS = ["RSV", "ADV", "FluB", "HPIV", "S.P", "C.P", "MP", "HRV", "FluA", "nCoV", "H.I", "CoV"]
CHANNEL_TARGETS = {
    "A|ATTO": "RSV",
    "A|FAM": "ADV",
    "A|HEX": "FluB",
    "A|ROX": "HPIV",
    "A|CY5": "S.P",
    "A|CY5.5": "C.P",
    "B|ATTO": "MP",
    "B|FAM": "HRV",
    "B|HEX": "FluA",
    "B|ROX": "nCoV",
    "B|CY5": "H.I",
    "B|CY5.5": "CoV",
}
EVIDENCE_FEATURES = [
    "ctDetected",
    "ctValue",
    "concentrationPresent",
    "logConcentration",
    "riskScore",
]


@dataclass(frozen=True)
class DatasetSummary:
    samples: int
    source_rows: int
    positive: int
    negative: int
    source_file_sha256: str
    label_source: str
    targets: list[str]
    evidence_features: list[str]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def locate_demo(path: str | None) -> Path:
    if path:
        return Path(path).resolve()
    expected = "\u9633\u6027\u7387\u9884\u8b66"
    candidates = [
        file
        for file in Path(__file__).resolve().parents[3].glob("*.xlsx")
        if not file.name.startswith("~$") and expected in file.name
    ]
    if len(candidates) != 1:
        raise ValueError("Pass --input; positive-rate warning workbook was not uniquely located")
    return candidates[0]


def parse_pairs(value: object) -> dict[str, str]:
    if value in (None, ""):
        return {}
    return {
        key.strip(): label.strip()
        for key, label in re.findall(r"\[\s*([^,\]]+)\s*,\s*([^\]]+)\]", str(value))
    }


def measurement(value: object) -> float | None:
    if value in (None, ""):
        return None
    parsed = float(value)
    return None if parsed <= -99 else parsed


def load_dataset(
    path: Path,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, DatasetSummary]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Order Detection Data"]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").replace(" ", "") for value in next(rows)]
    header_index = {value.upper(): index for index, value in enumerate(headers)}
    result_index = header_index["\u68c0\u6d4b\u7ed3\u679c"]

    evidence_rows: list[list[float]] = []
    target_rows: list[list[float]] = []
    labels: list[int] = []
    groups: list[int] = []
    source_rows = 0
    positive_label = "\u9633\u6027"
    for group, row in enumerate(rows):
        system_results = parse_pairs(row[result_index])
        if not system_results:
            continue
        source_rows += 1
        for path_key, target_code in CHANNEL_TARGETS.items():
            if target_code not in system_results:
                continue
            chamber, channel = path_key.split("|", 1)
            ct = measurement(
                row[header_index[f"{chamber}\u8154\u5ba4{channel}\u901a\u9053CT\u503c".upper()]]
            )
            concentration = measurement(
                row[header_index[f"{chamber}\u8154\u5ba4{channel}\u901a\u9053\u6d53\u5ea6".upper()]]
            )
            positive = system_results[target_code] == positive_label
            watch = (ct is not None and ct >= 35.0) or (
                concentration is not None and concentration <= 1_000.0
            )
            target_vector = [0.0] * len(TARGETS)
            target_vector[TARGETS.index(target_code)] = 1.0
            target_rows.append(target_vector)
            evidence_rows.append(
                [
                    1.0 if ct is not None else 0.0,
                    ct or 0.0,
                    1.0 if concentration is not None else 0.0,
                    float(np.log1p(max(concentration or 0.0, 0.0))),
                    1.0 if watch else 0.0,
                ]
            )
            labels.append(1 if positive else 0)
            groups.append(group)

    evidence = np.asarray(evidence_rows, dtype=np.float32)
    targets = np.asarray(target_rows, dtype=np.float32)
    target = np.asarray(labels, dtype=np.int64)
    group_ids = np.asarray(groups, dtype=np.int64)
    summary = DatasetSummary(
        samples=len(target),
        source_rows=source_rows,
        positive=int(target.sum()),
        negative=int((target == 0).sum()),
        source_file_sha256=sha256(path),
        label_source="WEAK_LABEL_FROM_HISTORICAL_SYSTEM_JUDGEMENT",
        targets=TARGETS,
        evidence_features=EVIDENCE_FEATURES,
    )
    return evidence, targets, target, group_ids, summary


def split_groups(features: np.ndarray, target: np.ndarray, groups: np.ndarray):
    splitter = GroupShuffleSplit(n_splits=20, test_size=0.2, random_state=RANDOM_STATE)
    for train_index, test_index in splitter.split(features, target, groups):
        if len(np.unique(target[train_index])) == 2 and len(np.unique(target[test_index])) == 2:
            return train_index, test_index
    raise ValueError("Could not create a grouped split containing both classes")


def evaluate(target: np.ndarray, predicted: np.ndarray) -> dict[str, object]:
    tn, fp, fn, tp = confusion_matrix(target, predicted, labels=[0, 1]).ravel()
    return {
        "accuracy": accuracy_score(target, predicted),
        "balancedAccuracy": balanced_accuracy_score(target, predicted),
        "sensitivity": tp / (tp + fn) if tp + fn else 0.0,
        "specificity": tn / (tn + fp) if tn + fp else 0.0,
        "confusionMatrix": {"tn": int(tn), "fp": int(fp), "fn": int(fn), "tp": int(tp)},
    }


def export_onnx(
    path: Path,
    scaler: StandardScaler,
    classifier: LogisticRegression,
) -> None:
    feature_count = len(TARGETS) + len(EVIDENCE_FEATURES)
    coefficient = np.zeros((feature_count, 4), dtype=np.float32)
    intercept = np.full(4, -20.0, dtype=np.float32)
    coefficient[:, 0] = -classifier.coef_[0].astype(np.float32) / 2
    coefficient[:, 1] = classifier.coef_[0].astype(np.float32) / 2
    intercept[0] = -float(classifier.intercept_[0]) / 2
    intercept[1] = float(classifier.intercept_[0]) / 2
    initializers = [
        numpy_helper.from_array(scaler.mean_.astype(np.float32), "mean"),
        numpy_helper.from_array(scaler.scale_.astype(np.float32), "scale"),
        numpy_helper.from_array(coefficient, "coefficient"),
        numpy_helper.from_array(intercept, "intercept"),
    ]
    nodes = [
        helper.make_node("Concat", ["target", "evidence"], ["combined"], axis=1),
        helper.make_node("Sub", ["combined", "mean"], ["centered"]),
        helper.make_node("Div", ["centered", "scale"], ["standardized"]),
        helper.make_node("MatMul", ["standardized", "coefficient"], ["linear"]),
        helper.make_node("Add", ["linear", "intercept"], ["logits"]),
    ]
    graph = helper.make_graph(
        nodes,
        "positive-rate-structured-classifier",
        [
            helper.make_tensor_value_info("target", TensorProto.FLOAT, [None, len(TARGETS)]),
            helper.make_tensor_value_info(
                "evidence", TensorProto.FLOAT, [None, len(EVIDENCE_FEATURES)]
            ),
        ],
        [helper.make_tensor_value_info("logits", TensorProto.FLOAT, [None, 4])],
        initializers,
    )
    model = helper.make_model(
        graph,
        producer_name="ai-audit-positive-rate-structured-training",
        opset_imports=[helper.make_opsetid("", 17)],
    )
    model.ir_version = 9
    onnx.checker.check_model(model)
    path.parent.mkdir(parents=True, exist_ok=True)
    onnx.save(model, path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input")
    parser.add_argument(
        "--output", default="artifacts/structured-classifier/0.1.0-warning-weak-label"
    )
    args = parser.parse_args()
    input_path = locate_demo(args.input)
    output_dir = Path(args.output).resolve()
    evidence, targets, labels, groups, summary = load_dataset(input_path)
    combined = np.concatenate([targets, evidence], axis=1)
    train_index, test_index = split_groups(combined, labels, groups)
    scaler = StandardScaler().fit(combined[train_index])
    classifier = LogisticRegression(
        class_weight="balanced", max_iter=5_000, random_state=RANDOM_STATE
    ).fit(scaler.transform(combined[train_index]), labels[train_index])
    expected = classifier.predict(scaler.transform(combined[test_index]))
    metrics = evaluate(labels[test_index], expected)
    metrics.update(
        {
            "modelCode": MODEL_CODE,
            "version": MODEL_VERSION,
            "intendedUse": "DEMO_STRUCTURED_AUXILIARY_ONLY",
            "dataset": asdict(summary),
            "trainSamples": int(len(train_index)),
            "testSamples": int(len(test_index)),
            "splitStrategy": "GROUPED_BY_DETECTION_ORDER",
            "preprocessing": "TARGET_ONE_HOT_PLUS_CT_CONCENTRATION_RISK_EVIDENCE_EMBEDDED_IN_ONNX",
            "metricWarning": "Agreement with historical system weak labels only; not clinical validation",
        }
    )
    model_path = output_dir / "model.onnx"
    export_onnx(model_path, scaler, classifier)
    session = ort.InferenceSession(str(model_path), providers=["CPUExecutionProvider"])
    logits = session.run(
        ["logits"],
        {
            "target": targets[test_index].astype(np.float32),
            "evidence": evidence[test_index].astype(np.float32),
        },
    )[0]
    if not np.array_equal(expected, logits.argmax(axis=1)):
        raise ValueError("ONNX predictions differ from the trained classifier")
    metrics["onnxParity"] = True
    metrics["onnxOutputShape"] = list(logits.shape)
    metrics["modelSha256"] = sha256(model_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "metrics.json").write_text(
        json.dumps(metrics, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(metrics, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
