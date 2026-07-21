from __future__ import annotations

import argparse
import ast
import hashlib
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path

import numpy as np
import onnx
import onnxruntime as ort
from onnx import TensorProto, helper, numpy_helper
from openpyxl import load_workbook
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, balanced_accuracy_score, confusion_matrix
from sklearn.model_selection import GroupShuffleSplit
from sklearn.preprocessing import StandardScaler


RANDOM_STATE = 42
MODEL_CODE = "curve-classifier"
MODEL_VERSION = "0.2.0-evidence-cold-start"
AUXILIARY_FEATURES = [
    "ctDetected",
    "ctValue",
    "concentrationPresent",
    "logConcentration",
    "riskScore",
]


@dataclass(frozen=True)
class DatasetSummary:
  samples: int
  positive: int
  negative: int
  points: int
  source_file_sha256: str
  label_source: str
  auxiliary_features: list[str]


def sha256(path: Path) -> str:
  digest = hashlib.sha256()
  with path.open("rb") as source:
    for block in iter(lambda: source.read(1024 * 1024), b""):
      digest.update(block)
  return digest.hexdigest()


def locate_demo(path: str | None) -> Path:
  if path:
    return Path(path).resolve()
  candidates = [
      file for file in Path(__file__).resolve().parents[3].glob("*.xlsx")
      if not file.name.startswith("~$") and file.stat().st_size > 300_000
  ]
  if len(candidates) != 1:
    raise ValueError("Pass --input; fluorescence demo workbook was not uniquely located")
  return candidates[0]


def parse_pairs(value: object) -> dict[str, str]:
  if value in (None, ""):
    return {}
  return {
      key.strip(): label.strip()
      for key, label in re.findall(r"\[\s*([^,\]]+)\s*,\s*([^\]]+)\]", str(value))
  }


def parse_mapping(value: object) -> dict[str, str]:
  if value in (None, ""):
    return {}
  result: dict[str, str] = {}
  for target, chamber, channel in re.findall(
      r"([^,{]+):\s*\[\s*([01AB])_([^\]]+)\]", str(value), flags=re.IGNORECASE
  ):
    normalized_chamber = "A" if chamber.upper() in {"0", "A"} else "B"
    normalized_channel = channel.replace(".", "").upper()
    result[f"{normalized_chamber}|{normalized_channel}"] = target.strip()
  return result


def measurement(value: object) -> float | None:
  if value in (None, ""):
    return None
  parsed = float(value)
  return None if parsed <= -99 else parsed


def load_weak_label_dataset(
    path: Path,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, DatasetSummary]:
  worksheet = load_workbook(path, read_only=True, data_only=True).active
  rows = worksheet.iter_rows(values_only=True)
  headers = [str(value or "").replace(" ", "") for value in next(rows)]
  header_index = {value.upper(): index for index, value in enumerate(headers)}
  raw_columns: list[tuple[str, int]] = []
  for index, header in enumerate(headers):
    matched = re.fullmatch(r"([AB])腔室原始数据_(.+)", header, flags=re.IGNORECASE)
    if matched:
      channel = matched.group(2).replace(".", "").upper()
      raw_columns.append((f"{matched.group(1).upper()}|{channel}", index))
  if not raw_columns:
    raise ValueError("No raw fluorescence curve columns found")
  result_index = header_index["检测结果"]
  mapping_index = header_index["病原和通道对应关系"]

  curves: list[list[float]] = []
  evidence_rows: list[list[float]] = []
  labels: list[int] = []
  groups: list[int] = []
  expected_points: int | None = None
  for group, row in enumerate(rows):
    mapping = parse_mapping(row[mapping_index])
    system_results = parse_pairs(row[result_index])
    for path_key, raw_index in raw_columns:
      target_code = mapping.get(path_key)
      if target_code is None or target_code not in system_results:
        continue
      if row[raw_index] in (None, ""):
        continue
      values = [float(value) for value in ast.literal_eval(str(row[raw_index]))]
      if not values:
        continue
      expected_points = expected_points or len(values)
      if len(values) != expected_points or not np.isfinite(values).all():
        continue
      chamber, channel = path_key.split("|", 1)
      excel_channel = "CY5.5" if channel == "CY55" else channel
      ct = measurement(row[header_index[f"{chamber}腔室{excel_channel}通道CT值".upper()]])
      concentration = measurement(
          row[header_index[f"{chamber}腔室{excel_channel}通道浓度".upper()]]
      )
      positive = system_results[target_code] == "阳性"
      watch = positive and (
          (ct is not None and ct >= 35.0)
          or (concentration is not None and concentration <= 1_000.0)
      )
      curves.append(values)
      evidence_rows.append([
          1.0 if ct is not None else 0.0,
          ct or 0.0,
          1.0 if concentration is not None else 0.0,
          float(np.log1p(max(concentration or 0.0, 0.0))),
          1.0 if watch else 0.0,
      ])
      labels.append(1 if positive else 0)
      groups.append(group)

  features = np.asarray(curves, dtype=np.float32)
  auxiliary = np.asarray(evidence_rows, dtype=np.float32)
  target = np.asarray(labels, dtype=np.int64)
  group_ids = np.asarray(groups, dtype=np.int64)
  summary = DatasetSummary(
      samples=len(target),
      positive=int(target.sum()),
      negative=int((target == 0).sum()),
      points=int(features.shape[1]),
      source_file_sha256=sha256(path),
      label_source="WEAK_LABEL_FROM_SYSTEM_JUDGEMENT",
      auxiliary_features=AUXILIARY_FEATURES,
  )
  return features, auxiliary, target, group_ids, summary


def split_groups(features: np.ndarray, target: np.ndarray, groups: np.ndarray):
  splitter = GroupShuffleSplit(n_splits=20, test_size=0.2, random_state=RANDOM_STATE)
  for train_index, test_index in splitter.split(features, target, groups):
    if len(np.unique(target[train_index])) == 2 and len(np.unique(target[test_index])) == 2:
      return train_index, test_index
  raise ValueError("Could not create a grouped split containing both classes")


def baseline_correct(features: np.ndarray) -> np.ndarray:
  return features - features[:, :10].mean(axis=1, keepdims=True)


def evaluate(target: np.ndarray, predicted: np.ndarray) -> dict[str, object]:
  tn, fp, fn, tp = confusion_matrix(target, predicted, labels=[0, 1]).ravel()
  return {
      "accuracy": accuracy_score(target, predicted),
      "balancedAccuracy": balanced_accuracy_score(target, predicted),
      "sensitivity": tp / (tp + fn) if tp + fn else 0.0,
      "specificity": tn / (tn + fp) if tn + fp else 0.0,
      "confusionMatrix": {"tn": int(tn), "fp": int(fp), "fn": int(fn), "tp": int(tp)},
  }


def export_onnx(path: Path, scaler: StandardScaler, classifier: LogisticRegression, points: int):
  coefficient = classifier.coef_.astype(np.float32).T
  intercept = classifier.intercept_.astype(np.float32)
  initializers = [
      numpy_helper.from_array(scaler.mean_.astype(np.float32), "mean"),
      numpy_helper.from_array(scaler.scale_.astype(np.float32), "scale"),
      numpy_helper.from_array(coefficient, "coefficient"),
      numpy_helper.from_array(intercept, "intercept"),
      numpy_helper.from_array(np.asarray([0.5], dtype=np.float32), "half"),
      numpy_helper.from_array(np.asarray([-20.0], dtype=np.float32), "inactive_logit"),
      numpy_helper.from_array(np.asarray([0], dtype=np.int64), "slice_starts"),
      numpy_helper.from_array(np.asarray([10], dtype=np.int64), "slice_ends"),
      numpy_helper.from_array(np.asarray([1], dtype=np.int64), "slice_axes"),
      numpy_helper.from_array(np.asarray([1], dtype=np.int64), "slice_steps"),
  ]
  nodes = [
      helper.make_node(
          "Slice",
          ["curve", "slice_starts", "slice_ends", "slice_axes", "slice_steps"],
          ["baseline_points"],
      ),
      helper.make_node("ReduceMean", ["baseline_points"], ["curve_baseline"], axes=[1], keepdims=1),
      helper.make_node("Sub", ["curve", "curve_baseline"], ["baseline_corrected"]),
      helper.make_node("Concat", ["baseline_corrected", "evidence"], ["combined"], axis=1),
      helper.make_node("Sub", ["combined", "mean"], ["centered"]),
      helper.make_node("Div", ["centered", "scale"], ["standardized"]),
      helper.make_node("MatMul", ["standardized", "coefficient"], ["linear"]),
      helper.make_node("Add", ["linear", "intercept"], ["score"]),
      helper.make_node("Mul", ["score", "half"], ["positive_logit"]),
      helper.make_node("Neg", ["positive_logit"], ["negative_logit"]),
      helper.make_node("Shape", ["score"], ["score_shape"]),
      helper.make_node("Expand", ["inactive_logit", "score_shape"], ["suspicious_logit"]),
      helper.make_node("Expand", ["inactive_logit", "score_shape"], ["invalid_logit"]),
      helper.make_node(
          "Concat",
          ["negative_logit", "positive_logit", "suspicious_logit", "invalid_logit"],
          ["logits"],
          axis=1,
      ),
  ]
  graph = helper.make_graph(
      nodes,
      "cold-start-curve-classifier",
      [
          helper.make_tensor_value_info("curve", TensorProto.FLOAT, [None, points]),
          helper.make_tensor_value_info(
              "evidence", TensorProto.FLOAT, [None, len(AUXILIARY_FEATURES)]
          ),
      ],
      [helper.make_tensor_value_info("logits", TensorProto.FLOAT, [None, 4])],
      initializers,
  )
  model = helper.make_model(
      graph,
      producer_name="ai-audit-cold-start-training",
      opset_imports=[helper.make_opsetid("", 17)],
  )
  model.ir_version = 9
  onnx.checker.check_model(model)
  path.parent.mkdir(parents=True, exist_ok=True)
  onnx.save(model, path)


def main():
  parser = argparse.ArgumentParser()
  parser.add_argument("--input")
  parser.add_argument("--output", default="artifacts/curve-classifier/0.2.0-evidence-cold-start")
  args = parser.parse_args()
  input_path = locate_demo(args.input)
  output_dir = Path(args.output).resolve()
  features, auxiliary, target, groups, summary = load_weak_label_dataset(input_path)
  train_index, test_index = split_groups(features, target, groups)
  corrected = baseline_correct(features)
  combined = np.concatenate([corrected, auxiliary], axis=1)

  scaler = StandardScaler().fit(combined[train_index])
  classifier = LogisticRegression(
      class_weight="balanced", max_iter=5_000, random_state=RANDOM_STATE
  ).fit(scaler.transform(combined[train_index]), target[train_index])
  expected = classifier.predict(scaler.transform(combined[test_index]))
  metrics = evaluate(
      target[test_index], expected
  )
  metrics.update({
      "modelCode": MODEL_CODE,
      "version": MODEL_VERSION,
      "intendedUse": "DEMO_COLD_START_ONLY",
      "dataset": asdict(summary),
      "trainSamples": int(len(train_index)),
      "testSamples": int(len(test_index)),
      "splitStrategy": "GROUPED_BY_INSTRUMENT_RUN",
      "preprocessing": "BASELINE_CORRECTION_PLUS_CT_CONCENTRATION_RISK_EVIDENCE_EMBEDDED_IN_ONNX",
      "metricWarning": "Agreement with system weak labels only; quantitative evidence is correlated with those labels and is not clinical validation",
  })

  model_path = output_dir / "model.onnx"
  export_onnx(model_path, scaler, classifier, summary.points)
  session = ort.InferenceSession(str(model_path), providers=["CPUExecutionProvider"])
  logits = session.run(
      ["logits"],
      {
          "curve": features[test_index].astype(np.float32),
          "evidence": auxiliary[test_index].astype(np.float32),
      },
  )[0]
  onnx_predicted = logits.argmax(axis=1)
  if not np.array_equal(expected, onnx_predicted):
    raise ValueError("ONNX predictions differ from the trained classifier")
  metrics["onnxParity"] = True
  metrics["onnxOutputShape"] = list(logits.shape)
  metrics["modelSha256"] = sha256(model_path)
  output_dir.mkdir(parents=True, exist_ok=True)
  (output_dir / "metrics.json").write_text(
      json.dumps(metrics, ensure_ascii=False, indent=2), encoding="utf-8"
  )
  print(json.dumps(metrics, ensure_ascii=False, indent=2))


if __name__ == "__main__":
  main()
