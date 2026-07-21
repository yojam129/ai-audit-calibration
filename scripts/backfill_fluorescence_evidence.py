from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

import mysql.connector
from openpyxl import load_workbook


def measurement(value: object) -> float | None:
    if value in (None, ""):
        return None
    parsed = float(value)
    return None if parsed <= -99 else parsed


def numeric_value(value: object, default: float = -100.0) -> float:
    if value in (None, ""):
        return default
    return float(value)


def mapping(value: object, default_chamber: str) -> dict[tuple[str, str], str]:
    result: dict[tuple[str, str], str] = {}
    for target, chamber, channel in re.findall(
        r"([^,{]+):\s*\[\s*(?:([01AB])_)?([^\]]+)\]", str(value), re.IGNORECASE
    ):
        chamber_code = default_chamber if not chamber else (
            "A" if chamber.upper() in {"0", "A"} else "B"
        )
        result[(chamber_code, channel.replace(".", "").upper())] = target.strip()
    return result


def system_results(value: object) -> dict[str, str]:
    return {
        target.strip(): label.strip()
        for target, label in re.findall(r"\[\s*([^,\]]+)\s*,\s*([^\]]+)\]", str(value))
    }


def active_chamber(row: tuple[object, ...], indexes: dict[str, int]) -> str:
    counts = {}
    for chamber in ("A", "B"):
        counts[chamber] = sum(
            row[index] not in (None, "")
            for header, index in indexes.items()
            if header.startswith(f"{chamber}腔室原始数据_")
        )
    if counts["A"] > 0 and counts["B"] == 0:
        return "A"
    if counts["B"] > 0 and counts["A"] == 0:
        return "B"
    return "A"


def run_number(instrument: object, started_at: object, module: object) -> str:
    started = started_at if isinstance(started_at, datetime) else datetime.fromisoformat(str(started_at))
    return f"{instrument}-{started:%Y%m%d%H%M%S}-{module if module not in (None, '') else 0}"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--host", required=True)
    parser.add_argument("--user", required=True)
    parser.add_argument("--password", required=True)
    parser.add_argument("--database", default="ai_audit_sample")
    args = parser.parse_args()

    worksheet = load_workbook(Path(args.input), read_only=True, data_only=True).active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").replace(" ", "") for value in next(rows)]
    indexes = {value.upper(): index for index, value in enumerate(headers)}
    connection = mysql.connector.connect(
        host=args.host, user=args.user, password=args.password, database=args.database
    )
    updated = 0
    updated_runs = 0
    try:
        cursor = connection.cursor()
        for row in rows:
            if row[indexes["仪器SN"]] in (None, "") or row[indexes["检测开始时间"]] in (None, ""):
                continue
            run_no = run_number(
                row[indexes["仪器SN"]], row[indexes["检测开始时间"]], row[indexes["模块位置"]]
            )
            instrument_type_value = row[indexes["仪器类型"]]
            instrument_type = (
                None if instrument_type_value in (None, "") else str(instrument_type_value)
            )
            module_position_value = row[indexes["模块位置"]]
            module_position = (
                None if module_position_value in (None, "") else str(module_position_value)
            )
            irc_a_raw = numeric_value(row[indexes["A腔室IRC通道CT值"]])
            irc_b_raw = numeric_value(row[indexes["B腔室IRC通道CT值"]])
            chamber_hint = active_chamber(row, indexes)
            target_mapping = mapping(
                row[indexes["病原和通道对应关系"]], chamber_hint
            )
            active_chambers = sorted({chamber for chamber, _ in target_mapping})
            irc_by_chamber = {"A": irc_a_raw, "B": irc_b_raw}
            qc_passed = bool(active_chambers) and all(
                irc_by_chamber[chamber] > -99 for chamber in active_chambers
            )
            qc_evidence_json = json.dumps(
                {
                    "A_IRC_CT": irc_a_raw,
                    "B_IRC_CT": irc_b_raw,
                    "activeChambers": active_chambers,
                    "logic": "活动腔室 IRC Ct 均为有效值（大于 -99）时质控通过",
                },
                ensure_ascii=False,
            )
            cursor.execute(
                """
                UPDATE instrument_run
                SET module_position = %s,
                    instrument_type = %s,
                    qc_status = %s,
                    qc_evidence_json = %s
                WHERE run_no = %s
                """,
                (
                    module_position,
                    instrument_type,
                    "PASS" if qc_passed else "INVALID",
                    qc_evidence_json,
                    run_no,
                ),
            )
            updated_runs += cursor.rowcount
            unit = str(row[indexes["浓度单位"]] or "Copies/mL")
            labels = system_results(row[indexes["检测结果"]])
            for (chamber, channel), target in target_mapping.items():
                excel_channel = "CY5.5" if channel == "CY55" else channel
                concentration = measurement(
                    row[indexes[f"{chamber}腔室{excel_channel}通道浓度".upper()]]
                )
                ct = measurement(row[indexes[f"{chamber}腔室{excel_channel}通道CT值".upper()]])
                flags = []
                positive = labels.get(target) == "阳性"
                if positive and ct is not None and ct >= 35:
                    flags.append("BORDERLINE_CT")
                if positive and concentration is not None and concentration <= 1_000:
                    flags.append("LOW_CONCENTRATION")
                risk_level = "WATCH" if flags else "NORMAL"
                cursor.execute(
                    """
                    UPDATE target_judgement t
                    JOIN instrument_run r ON r.id = t.run_id
                    SET t.channel_code = COALESCE(t.channel_code, %s),
                        t.concentration_value = %s,
                        t.concentration_unit = COALESCE(t.concentration_unit, %s),
                        t.risk_level = %s,
                        t.risk_flags = %s
                    WHERE r.run_no = %s AND t.chamber = %s
                      AND REPLACE(COALESCE(t.channel_code, t.target_code), '.', '') = %s
                      AND t.concentration_value IS NULL
                    """,
                    (channel, concentration, unit, risk_level, ",".join(flags) or None,
                     run_no, chamber, channel),
                )
                updated += cursor.rowcount
        connection.commit()
    finally:
        connection.close()
    print(f"Updated instrument runs: {updated_runs}")
    print(f"Updated target evidence rows: {updated}")


if __name__ == "__main__":
    main()
